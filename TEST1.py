from openpyxl import Workbook

# 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
wb = Workbook()

ws = wb.get_active_sheet()
print(ws.title)

ws.title = 'New Title'  # 设置worksheet的标题

# 设置单元格的值
ws.cell('D3').value = 4
ws.cell(row=3, column=1).value = 6

new_ws = wb.create_sheet(title='new_sheet')
for row in range(100):
    for col in range(10):
        new_ws.cell(row=row, column=col).value = row + col

# 最后一定要保存！
wb.save(filename='new_file.xlsx')
